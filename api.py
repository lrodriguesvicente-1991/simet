# =====================================================================
# ARQUIVO: api.py
# DESCRIÇÃO: API RESTful (FastAPI) Segura com JWT e Leitura de Pipe
# =====================================================================

import io
import json
import os
import sys
import time
from collections import deque
import psutil
import subprocess
import threading
import jwt
from datetime import datetime, timedelta, timezone
from fastapi import FastAPI, HTTPException, Depends, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.gzip import GZipMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
import pandas as pd
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

from database.connection import obter_conexao
from database.engine import contagem_saude_fila
from database import usuarios as usuarios_db
from robots._controle import limpar_parada, marcar_parada

load_dotenv()

app = FastAPI(title="SIMET API", version="2.0")

# =====================================================================
# CONFIGURAÇÃO DE CORS SEGURA
# =====================================================================
origins = os.getenv("CORS_ALLOWED_ORIGINS", "http://localhost:3000").split(",")

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# GZip reduz payload JSON do /api/dados em ~10x (texto numerico repetitivo)
app.add_middleware(GZipMiddleware, minimum_size=500)

# =====================================================================
# SISTEMA DE AUTENTICAÇÃO JWT (Seguro e via .ENV)
# =====================================================================
SECRET_KEY = os.getenv("JWT_SECRET_KEY")
if not SECRET_KEY:
    raise RuntimeError("JWT_SECRET_KEY não definida no .env — impossível iniciar a API com segurança.")
ALGORITHM = "HS256"
security = HTTPBearer()

class LoginRequest(BaseModel):
    usuario: str
    senha: str

try:
    usuarios_db.garantir_admin_seed()
except Exception as _seed_err:
    print(f"[SEED WARN] Falha ao garantir admin inicial: {_seed_err}", flush=True)


@app.post("/api/login")
def fazer_login(req: LoginRequest):
    user = usuarios_db.buscar_por_username(req.usuario)
    if not user or not user.get("ativo") or not usuarios_db.verificar_senha(req.senha, user["senha_hash"]):
        return {"sucesso": False, "mensagem": "Credenciais invalidas"}

    payload = {
        "sub": user["username"],
        "nivel": user["nivel"],
        "uid": user["id"],
        "exp": datetime.now(timezone.utc) + timedelta(hours=24),
    }
    token = jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)
    return {"sucesso": True, "token": token, "nivel": user["nivel"], "usuario": user["username"]}


def verificar_token(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        return {
            "username": payload["sub"],
            "nivel": int(payload.get("nivel", 3)),
            "id": payload.get("uid"),
        }
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token invalido")


def exigir_nivel(nivel_maximo: int):
    """Depende de verificar_token e barra usuarios com nivel MAIOR que o permitido.
    Lembre-se: 0=Admin (mais poder), 3=Visualizador (menos poder)."""
    def _dep(user: dict = Depends(verificar_token)) -> dict:
        if user["nivel"] > nivel_maximo:
            raise HTTPException(status_code=403, detail="Permissao insuficiente para esta acao.")
        return user
    return _dep

# =====================================================================
# GERENCIAMENTO DE ESTADO E PROCESSOS (Com Threading)
# =====================================================================
CB_MARKER = "[CIRCUIT_BREAKER]"

estado_robo = {
    "rodando": False,
    "processo": None, # Guarda o objeto Popen
    "logs": deque(maxlen=500),
    "circuit_breaker": False,
    "circuit_breaker_msg": None,
    "parando": False,  # True entre o pedido gracioso e o termino dos workers
}

def ler_logs_robo(proc):
    """Lê os logs do stdout em tempo real numa Thread separada (Evita block do OS)"""
    for linha in iter(proc.stdout.readline, ''):
        if linha:
            linha_strip = linha.strip()
            estado_robo["logs"].append(linha_strip)
            if CB_MARKER in linha_strip:
                estado_robo["circuit_breaker"] = True
                estado_robo["circuit_breaker_msg"] = linha_strip
    proc.stdout.close()

class RoboRequest(BaseModel):
    task: str
    workers: int = 1
    limit: int = 0
    headless: bool = True
    delay_extra: float = 0.0

@app.post("/api/robo/iniciar")
def iniciar_robo(req: RoboRequest, user: dict = Depends(exigir_nivel(1))):
    if req.task not in ("lfp", "eci", "aci", "full", "test"):
        raise HTTPException(status_code=400, detail=f"Tarefa invalida: {req.task}")

    if estado_robo["rodando"]:
        # Double check se não morreu silênciosamente
        if estado_robo["processo"] and estado_robo["processo"].poll() is None:
            return {"sucesso": False, "mensagem": "Um processo já está em andamento."}

    try:
        env_utf8 = os.environ.copy()
        env_utf8["PYTHONIOENCODING"] = "utf-8"
        env_utf8["SIMET_HEADLESS"] = "1" if req.headless else "0"
        env_utf8["SIMET_DELAY_EXTRA_S"] = str(max(0.0, req.delay_extra))

        comando = [sys.executable, "-u", "main.py", "--task", req.task, "--workers", str(req.workers)]
        if req.limit > 0: comando.extend(["--limit", str(req.limit)])
        
        processo = subprocess.Popen(
            comando, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, 
            text=True, bufsize=1, env=env_utf8, encoding='utf-8'
        )
        
        estado_robo["rodando"] = True
        estado_robo["processo"] = processo
        estado_robo["logs"] = deque([f"[SYSTEM] Iniciando: {req.task.upper()}..."], maxlen=500)
        estado_robo["circuit_breaker"] = False
        estado_robo["circuit_breaker_msg"] = None
        estado_robo["parando"] = False
        limpar_parada()
        
        # Dispara a thread para não bloquear o pipe!
        t = threading.Thread(target=ler_logs_robo, args=(processo,), daemon=True)
        t.start()
        
        return {"sucesso": True, "mensagem": "Robôs iniciados!"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao iniciar robô: {e}")

@app.post("/api/robo/parar")
def parar_robo(force: bool = False, user: dict = Depends(exigir_nivel(1))):
    """Parada graciosa por padrao: sinaliza aos workers para encerrarem apos
    o anuncio atual. force=true mata a arvore de processos na hora (fallback)."""
    if not estado_robo["rodando"] or not estado_robo["processo"]:
        return {"sucesso": False, "mensagem": "Nenhum robo rodando no momento."}

    if not force:
        marcar_parada()
        estado_robo["parando"] = True
        estado_robo["logs"].append("[SYSTEM] Parando — aguardando os robôs terminarem o anúncio atual...")
        return {"sucesso": True, "mensagem": "Parando...", "parando": True}

    # force=True -> taskkill imediato (mantem o comportamento antigo como fallback)
    pid = estado_robo["processo"].pid
    mortos = 0

    if sys.platform == "win32":
        try:
            subprocess.run(
                ["taskkill", "/F", "/T", "/PID", str(pid)],
                capture_output=True, timeout=10, check=False,
            )
            mortos = 1
        except Exception as e:
            estado_robo["logs"].append(f"[SYSTEM] taskkill falhou: {e}")

    try:
        parent = psutil.Process(pid)
        filhos = parent.children(recursive=True)
        for child in filhos:
            try: child.kill()
            except psutil.NoSuchProcess: pass
        try: parent.kill()
        except psutil.NoSuchProcess: pass
        psutil.wait_procs([parent] + filhos, timeout=3)
        mortos = 1
    except psutil.NoSuchProcess:
        mortos = 1
    except Exception as e:
        estado_robo["logs"].append(f"[SYSTEM] psutil falhou: {e}")

    estado_robo["rodando"] = False
    estado_robo["processo"] = None
    estado_robo["parando"] = False
    limpar_parada()
    estado_robo["logs"].append("[SYSTEM] Operação interrompida pelo usuário.")

    return {"sucesso": bool(mortos), "mensagem": "Robos parados a forca." if mortos else "Falha ao matar processo."}

@app.get("/api/robo/status")
def status_robo(user: dict = Depends(exigir_nivel(2))):
    # Checagem de vida: O processo morreu naturalmente (terminou o limite ou parada graciosa)?
    if estado_robo["rodando"] and estado_robo["processo"]:
        if estado_robo["processo"].poll() is not None:
            estado_robo["rodando"] = False
            estado_robo["processo"] = None
            if estado_robo["parando"]:
                estado_robo["logs"].append("[SYSTEM] Operação encerrada.")
                estado_robo["parando"] = False
            else:
                estado_robo["logs"].append("[SYSTEM] Operação concluída.")
            limpar_parada()

    logs = estado_robo["logs"]
    ultimas = list(logs)[-50:] if logs else []
    return {
        "rodando": estado_robo["rodando"],
        "logs_recentes": ultimas,
        "circuit_breaker": estado_robo["circuit_breaker"],
        "circuit_breaker_msg": estado_robo["circuit_breaker_msg"],
        "parando": estado_robo["parando"],
    }

@app.post("/api/robo/reconhecer-alarme")
def reconhecer_alarme(user: dict = Depends(exigir_nivel(1))):
    """Administrador reconhece o alarme do circuit breaker e limpa a flag."""
    estado_robo["circuit_breaker"] = False
    estado_robo["circuit_breaker_msg"] = None
    return {"sucesso": True}

# =====================================================================
# ROTAS DE DASHBOARD (Requerem Token)
# =====================================================================
# Cache in-memory do payload do Observatorio. Invalida em /api/sincronizar.
# A view nao muda fora dos refreshes da matview, entao 10min e seguro.
_DADOS_CACHE: dict = {"timestamp": 0.0, "payload": None}
_DADOS_TTL_S = 600

_QUERY_DADOS = """
SELECT
    m.mun_nome AS municipio,
    uf.unf_sigla AS estado,
    CASE uf.unf_reg_id
        WHEN 1 THEN 'Norte' WHEN 2 THEN 'Nordeste'
        WHEN 3 THEN 'Sudeste' WHEN 4 THEN 'Sul'
        WHEN 5 THEN 'Centro-Oeste' ELSE 'Desconhecida'
    END AS regiao,
    calc.categoria_tamanho, calc.total_anuncios_reais,
    calc.mediana_geral, calc.mediana_agricola, calc.mediana_pecuaria,
    calc.mediana_floresta_plantada, calc.mediana_floresta_nativa,
    calc.n_agricola, calc.n_pecuaria, calc.n_floresta_plantada, calc.n_floresta_nativa,
    calc.media_geral, calc.media_agricola, calc.media_pecuaria,
    calc.media_floresta_plantada, calc.media_floresta_nativa,
    calc.desvio_padrao, calc.coef_dispersao_pct,
    mr.mre_codigo AS mercado_regional_codigo,
    mr.mre_nome   AS mercado_regional_nome,
    c.lat, c.lon
FROM public.mv_estatisticas_simet calc
JOIN public.smt_municipio m ON calc.anc_mun_id = m.mun_id
JOIN public.smt_unidade_federativa uf ON m.mun_unf_id = uf.unf_id
JOIN public.mv_centroide_municipio c ON c.mun_cod = m.mun_cod
LEFT JOIN public.smt_mercado_regional mr ON mr.mre_id = m.mun_mre_id
"""


def _invalidar_cache_dados() -> None:
    _DADOS_CACHE["timestamp"] = 0.0
    _DADOS_CACHE["payload"] = None


def _carregar_dados_observatorio() -> dict:
    """Executa a query e devolve o payload pronto para serializacao.
    Usa cursor direto (sem pandas) — em payloads grandes economiza ~1-2s
    de overhead de DataFrame -> dict."""
    t0 = time.time()
    conn = obter_conexao()
    try:
        cur = conn.cursor()
        cur.execute(_QUERY_DADOS)
        cols = [d[0] for d in cur.description]
        rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()
    t_query = time.time() - t0

    # Conversao manual: list[dict] com floats nativos. Mais rapido que
    # pandas.to_dict() em conjuntos de >5k linhas.
    dados = [dict(zip(cols, row)) for row in rows]
    payload = {"sucesso": True, "dados": dados}

    _DADOS_CACHE["payload"] = payload
    _DADOS_CACHE["timestamp"] = time.time()
    print(f"[DADOS] cache aquecido: {len(dados)} linhas, query {t_query:.2f}s, total {time.time()-t0:.2f}s", flush=True)
    return payload


def _preaquecer_cache_em_thread() -> None:
    def _alvo():
        try:
            _carregar_dados_observatorio()
        except Exception as e:
            print(f"[DADOS] pre-aquecimento falhou: {e}", flush=True)
    threading.Thread(target=_alvo, daemon=True).start()


@app.on_event("startup")
def _on_startup_preaquecer():
    """Dispara o carregamento do payload em background no boot da API,
    pra primeira chamada do Observatorio ja vir do cache."""
    _preaquecer_cache_em_thread()


@app.get("/api/dados")
def get_dashboard_data(user: dict = Depends(verificar_token)):
    """Busca os dados mastigados da View e converte para JSON para o React."""
    agora = time.time()
    if _DADOS_CACHE["payload"] is not None and (agora - _DADOS_CACHE["timestamp"]) < _DADOS_TTL_S:
        return _DADOS_CACHE["payload"]
    try:
        return _carregar_dados_observatorio()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/mercados-regionais")
def listar_mercados_regionais(user: dict = Depends(verificar_token)):
    """Lista os mercados regionais (tipo smt_mercado_regional) com a UF inferida do codigo."""
    try:
        conn = obter_conexao()
        cur = conn.cursor()
        cur.execute("""
            SELECT DISTINCT mr.mre_codigo, mr.mre_nome, u.unf_sigla
            FROM public.smt_mercado_regional mr
            LEFT JOIN public.smt_municipio m ON m.mun_mre_id = mr.mre_id
            LEFT JOIN public.smt_unidade_federativa u ON u.unf_id = m.mun_unf_id
            ORDER BY u.unf_sigla NULLS LAST, mr.mre_nome
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return {
            "sucesso": True,
            "mercados": [
                {"codigo": r[0], "nome": r[1], "uf": (r[2] or "").strip() or None}
                for r in rows
            ],
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/geom")
def get_geom(tipo: str, nome: str | None = None, uf: str | None = None,
             user: dict = Depends(verificar_token)):
    """Retorna GeoJSON da feature selecionada (regiao | estado | municipio)."""
    try:
        conn = obter_conexao()
        cur = conn.cursor()

        if tipo == "regiao":
            if not nome: raise HTTPException(400, "Parametro 'nome' obrigatorio")
            cur.execute("""
                SELECT ST_AsGeoJSON(ST_SimplifyPreserveTopology(mlr_geom::geometry, 0.01))
                FROM public.smt_malha_regiao WHERE lower(mlr_nm_regiao) = lower(%s)
            """, (nome,))
        elif tipo == "estado":
            if not uf: raise HTTPException(400, "Parametro 'uf' obrigatorio")
            cur.execute("""
                SELECT ST_AsGeoJSON(ST_SimplifyPreserveTopology(mlu_geom::geometry, 0.01))
                FROM public.smt_malha_uf WHERE mlu_sigla_uf = %s
            """, (uf,))
        elif tipo == "municipio":
            if not (nome and uf): raise HTTPException(400, "Parametros 'nome' e 'uf' obrigatorios")
            cur.execute("""
                SELECT ST_AsGeoJSON(ST_SimplifyPreserveTopology(mlm_geom::geometry, 0.001))
                FROM public.smt_malha_municipal
                WHERE mlm_nm_mun = %s AND mlm_sigla_uf = %s
                LIMIT 1
            """, (nome, uf))
        else:
            raise HTTPException(400, "tipo deve ser: regiao | estado | municipio")

        row = cur.fetchone()
        cur.close()
        conn.close()

        if not row or not row[0]:
            return {"sucesso": False, "geom": None}

        return {"sucesso": True, "geom": json.loads(row[0])}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/sincronizar")
def sincronizar_view(user: dict = Depends(exigir_nivel(1))):
    """Atualiza as Materialized Views (estatisticas + centroides)."""
    try:
        conn = obter_conexao()
        conn.autocommit = True
        cur = conn.cursor()
        cur.execute("REFRESH MATERIALIZED VIEW CONCURRENTLY public.mv_estatisticas_simet;")
        cur.execute("REFRESH MATERIALIZED VIEW CONCURRENTLY public.mv_centroide_municipio;")
        cur.close()
        conn.close()
        _invalidar_cache_dados()
        # Reaquecimento sincrono: a proxima chamada do front a /api/dados ja
        # encontra o cache fresco, evitando que o usuario espere a query de
        # novo apos o sincronizar.
        try:
            _carregar_dados_observatorio()
        except Exception as e:
            print(f"[SINCRONIZAR] reaquecimento falhou: {e}", flush=True)
        return {"sucesso": True, "mensagem": "Base de dados sincronizada com sucesso!"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# =====================================================================
# EXPORTACAO DE RELATORIOS (XLSX e PDF)
# =====================================================================
def _buscar_dados_relatorio(regiao=None, estado=None, categoria=None,
                            municipio=None, mercado_regional=None):
    """Busca a view aplicando os mesmos filtros do Observatorio (SQL-side)."""
    wheres: list[str] = ["v.geom_municipio IS NOT NULL"]
    params: list = []
    if regiao and regiao.lower() not in ("", "todas"):
        wheres.append("v.regiao = %s"); params.append(regiao)
    if estado and estado.lower() not in ("", "todos"):
        wheres.append("v.estado = %s"); params.append(estado)
    if categoria and categoria.lower() not in ("", "todos"):
        cats = [c.strip() for c in categoria.split(",") if c.strip()]
        if len(cats) == 1:
            wheres.append("v.categoria_tamanho = %s"); params.append(cats[0])
        elif len(cats) > 1:
            placeholders = ", ".join(["%s"] * len(cats))
            wheres.append(f"v.categoria_tamanho IN ({placeholders})")
            params.extend(cats)
    if municipio and municipio.lower() not in ("", "todos"):
        wheres.append("v.municipio = %s"); params.append(municipio)
    if mercado_regional and mercado_regional.lower() not in ("", "todos"):
        wheres.append("mr.mre_codigo = %s"); params.append(mercado_regional)

    query = """
    SELECT v.municipio, v.estado, v.regiao, v.categoria_tamanho, v.total_anuncios_reais,
           v.mediana_geral, v.mediana_agricola, v.mediana_pecuaria,
           v.mediana_floresta_plantada, v.mediana_floresta_nativa,
           v.n_agricola, v.n_pecuaria, v.n_floresta_plantada, v.n_floresta_nativa,
           v.media_geral, v.media_agricola, v.media_pecuaria,
           v.media_floresta_plantada, v.media_floresta_nativa,
           v.desvio_padrao, v.coef_dispersao_pct,
           mr.mre_codigo AS mercado_regional_codigo,
           mr.mre_nome   AS mercado_regional_nome
    FROM public.vw_media_mercado_terras v
    LEFT JOIN public.smt_unidade_federativa u ON u.unf_sigla = v.estado
    LEFT JOIN public.smt_municipio m
           ON m.mun_nome = v.municipio AND m.mun_unf_id = u.unf_id
    LEFT JOIN public.smt_mercado_regional mr ON mr.mre_id = m.mun_mre_id
    WHERE """ + " AND ".join(wheres)

    conn = obter_conexao()
    try:
        df = pd.read_sql(query, conn, params=params or None)
    finally:
        conn.close()
    return df.replace({pd.NA: None, float('nan'): None})


@app.get("/api/relatorio/xlsx")
def exportar_xlsx(regiao: str | None = None, estado: str | None = None,
                  categoria: str | None = None, municipio: str | None = None,
                  mercado_regional: str | None = None,
                  user: dict = Depends(verificar_token)):
    """Exporta os dados filtrados em XLSX (sem limite de linhas)."""
    try:
        df = _buscar_dados_relatorio(regiao, estado, categoria, municipio, mercado_regional)

        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="SIMET")
            ws = writer.sheets["SIMET"]
            for col_idx, col_name in enumerate(df.columns, start=1):
                largura = max(12, min(30, len(str(col_name)) + 2))
                ws.column_dimensions[get_column_letter(col_idx)].width = largura

        buffer.seek(0)
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        nome = f"simet_{ts}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{nome}"'},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


_FAVICON_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             "frontend_react", "public", "favicon.png")


def _desenhar_logo_incra(c, x, y, tamanho=36):
    """Desenha o logotipo do INCRA a partir do favicon.png.
    Se o arquivo nao estiver disponivel, cai para o logotipo vetorial."""
    from reportlab.lib.colors import HexColor
    from reportlab.lib.utils import ImageReader

    if os.path.exists(_FAVICON_PATH):
        try:
            c.drawImage(ImageReader(_FAVICON_PATH), x, y,
                        width=tamanho, height=tamanho,
                        mask='auto', preserveAspectRatio=True)
            return
        except Exception:
            pass  # fallback vetorial abaixo

    verde = HexColor("#6FA030")
    t = tamanho
    c.setFillColor(verde)
    q = t * 0.18
    gap = t * 0.02
    ox = x + t * 0.34
    oy = y + t * 0.55
    for i in range(2):
        for j in range(2):
            c.rect(ox + i * (q + gap), oy + (1 - j) * (q + gap), q, q, fill=1, stroke=0)
    c.setFillColor(verde)
    p = c.beginPath(); p.moveTo(x, y + t * 0.05)
    p.lineTo(x + t * 0.44, y + t * 0.12); p.lineTo(x + t * 0.46, y + t * 0.55)
    p.lineTo(x, y + t * 0.5); p.close(); c.drawPath(p, fill=1, stroke=0)
    p = c.beginPath(); p.moveTo(x + t * 0.54, y + t * 0.12)
    p.lineTo(x + t, y + t * 0.05); p.lineTo(x + t, y + t * 0.5)
    p.lineTo(x + t * 0.52, y + t * 0.55); p.close(); c.drawPath(p, fill=1, stroke=0)


def _fmt_brl(v):
    if v is None or (isinstance(v, float) and (pd.isna(v) or not v)):
        return "—"
    try:
        return f"R$ {float(v):,.0f}".replace(",", ".")
    except Exception:
        return "—"


@app.get("/api/relatorio/pdf")
def exportar_pdf(estado: str, categoria: str | None = None,
                 mercado_regional: str | None = None,
                 user: dict = Depends(verificar_token)):
    """Exporta relatorio PDF para uma UF especifica."""
    if not estado or estado.lower() == "todos":
        raise HTTPException(status_code=400, detail="Parametro 'estado' (UF) e obrigatorio no PDF.")

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.colors import HexColor, black
        from reportlab.pdfgen import canvas
        from reportlab.platypus import Table, TableStyle
        from reportlab.lib import colors as rl_colors

        df = _buscar_dados_relatorio(regiao=None, estado=estado,
                                     categoria=categoria, mercado_regional=mercado_regional)
        df = df.sort_values(["municipio", "categoria_tamanho"], na_position="last")

        mercado_nome = ""
        if mercado_regional and mercado_regional.lower() != "todos" and not df.empty:
            nome_col = df["mercado_regional_nome"].dropna()
            if not nome_col.empty:
                mercado_nome = str(nome_col.iloc[0])

        buffer = io.BytesIO()
        largura, altura = landscape(A4)
        c = canvas.Canvas(buffer, pagesize=landscape(A4))

        def cabecalho():
            _desenhar_logo_incra(c, 36, altura - 60, tamanho=40)
            c.setFillColor(HexColor("#3D6A1C"))
            c.setFont("Helvetica-Bold", 18)
            c.drawString(90, altura - 42, "SIMET · Observatorio de Mercado de Terras")
            c.setFillColor(black)
            c.setFont("Helvetica", 10)
            sub = f"UF: {estado}"
            if mercado_nome:
                sub += f" · Mercado Regional: {mercado_nome}"
            if categoria and categoria.lower() != "todos":
                cats_txt = ", ".join(c.strip() for c in categoria.split(",") if c.strip())
                if cats_txt:
                    sub += f" · Categoria: {cats_txt}"
            sub += f" · Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            c.drawString(90, altura - 58, sub)
            c.setStrokeColor(HexColor("#6FA030"))
            c.setLineWidth(1.2)
            c.line(36, altura - 70, largura - 36, altura - 70)

        def rodape(pagina, total):
            c.setFont("Helvetica-Oblique", 8)
            c.setFillColor(HexColor("#888888"))
            c.drawString(36, 20, "SIMET · Sistema de Inteligencia de Mercado de Terras · INCRA")
            c.drawRightString(largura - 36, 20, f"Pagina {pagina}/{total}")
            c.setFillColor(black)

        if df.empty:
            cabecalho()
            c.setFont("Helvetica", 12)
            c.drawString(40, altura - 100, "Nenhum dado encontrado para os filtros selecionados.")
            rodape(1, 1)
            c.save()
        else:
            cabecalhos = ["Municipio", "Mercado Regional", "Categoria", "Mediana Geral", "Amostras", "Media Geral", "Disp. %"]
            linhas = [cabecalhos]
            for _, r in df.iterrows():
                linhas.append([
                    str(r["municipio"]),
                    str(r.get("mercado_regional_nome") or "—"),
                    str(r["categoria_tamanho"]),
                    _fmt_brl(r.get("mediana_geral")),
                    str(int(r.get("total_anuncios_reais") or 0)),
                    _fmt_brl(r.get("media_geral")),
                    f"{float(r['coef_dispersao_pct']):.1f}%" if r.get("coef_dispersao_pct") is not None else "—",
                ])

            chunk = 22
            total_pags = max(1, (len(linhas) - 1 + chunk - 1) // chunk)
            for p in range(total_pags):
                cabecalho()
                inicio = 1 + p * chunk
                fim = min(len(linhas), inicio + chunk)
                trecho = [cabecalhos] + linhas[inicio:fim]
                tbl = Table(trecho, colWidths=[120, 140, 110, 90, 60, 90, 55])
                tbl.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), HexColor("#3D6A1C")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.whitesmoke, rl_colors.white]),
                    ("GRID", (0, 0), (-1, -1), 0.25, HexColor("#CCCCCC")),
                    ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ]))
                w, h = tbl.wrapOn(c, largura - 72, altura - 140)
                tbl.drawOn(c, 36, altura - 80 - h)
                rodape(p + 1, total_pags)
                if p + 1 < total_pags:
                    c.showPage()
            c.save()

        buffer.seek(0)
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        nome = f"simet_{estado}_{ts}.pdf"
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{nome}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# =====================================================================
# RELATORIO DE GRANDES PROPRIEDADES (Layout INCRA/UFF — fazendas >= 50 ha)
# =====================================================================
_COLUNAS_GRANDES = [
    "Região", "Mercado Regional de Terras", "UF", "COD. IBGE", "Município",
    "Mediana Geral", "Média Geral", "CV Geral",
    "Mediana Agrícola", "Média Agrícola", "CV Agrícola",
    "Mediana Pecuária", "Média Pecuária", "CV Pecuária",
    "Mediana Floresta Plantada", "Média Floresta Plantada", "CV Floresta Plantada",
    "Mediana Vegetação Nativa", "Média Vegetação Nativa", "CV Vegetação Nativa",
]


def _buscar_dados_grandes(regiao=None, estado=None, municipio=None, mercado_regional=None):
    """Busca a matview de fazendas (>= 50 ha) no layout INCRA/UFF."""
    wheres: list[str] = []
    params: list = []
    if regiao and regiao.lower() not in ("", "todas"):
        wheres.append("v.regiao = %s"); params.append(regiao)
    if estado and estado.lower() not in ("", "todos"):
        wheres.append("v.uf = %s"); params.append(estado)
    if municipio and municipio.lower() not in ("", "todos"):
        wheres.append("v.municipio = %s"); params.append(municipio)
    if mercado_regional and mercado_regional.lower() not in ("", "todos"):
        wheres.append("mr.mre_codigo = %s"); params.append(mercado_regional)

    where_sql = ("WHERE " + " AND ".join(wheres)) if wheres else ""

    query = f"""
    SELECT v.regiao,
           v.mercado_regional,
           v.uf,
           v.mun_cod,
           v.municipio,
           v.mediana_geral, v.media_geral, v.coef_dispersao_geral,
           v.mediana_agricola, v.media_agricola, v.coef_dispersao_agricola,
           v.mediana_pecuaria, v.media_pecuaria, v.coef_dispersao_pecuaria,
           v.mediana_floresta_plantada, v.media_floresta_plantada, v.coef_dispersao_floresta_plantada,
           v.mediana_vegetacao_nativa, v.media_vegetacao_nativa, v.coef_dispersao_vegetacao_nativa
    FROM public.mv_media_municipio_fazendas v
    LEFT JOIN public.smt_unidade_federativa uf ON uf.unf_sigla = v.uf
    LEFT JOIN public.smt_municipio m
           ON m.mun_cod = v.mun_cod
    LEFT JOIN public.smt_mercado_regional mr ON mr.mre_id = m.mun_mre_id
    {where_sql}
    ORDER BY v.regiao, v.uf, v.municipio
    """

    conn = obter_conexao()
    try:
        df = pd.read_sql(query, conn, params=params or None)
    finally:
        conn.close()
    return df.replace({pd.NA: None, float('nan'): None})


def _fmt_brl_num(v):
    """Formata numero como '99.999.999,99' (sem prefixo R$)."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    try:
        return f"{float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return ""


def _fmt_pct(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    try:
        return f"{float(v):,.2f} %".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return ""


@app.get("/api/relatorio/grandes-propriedades/xlsx")
def exportar_grandes_xlsx(regiao: str | None = None, estado: str | None = None,
                          municipio: str | None = None, mercado_regional: str | None = None,
                          user: dict = Depends(verificar_token)):
    """Relatorio de Grandes Propriedades (fazendas >= 50 ha) no layout INCRA/UFF."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter as _gcl

        df = _buscar_dados_grandes(regiao, estado, municipio, mercado_regional)

        wb = Workbook()
        ws = wb.active
        ws.title = "Grandes Propriedades"

        verde = PatternFill(start_color="3D6A1C", end_color="3D6A1C", fill_type="solid")
        verde_claro = PatternFill(start_color="6FA030", end_color="6FA030", fill_type="solid")
        branco_bold = Font(bold=True, color="FFFFFF", size=10)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(border_style="thin", color="DDDDDD")
        borda = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Titulo
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_COLUNAS_GRANDES))
        cel = ws.cell(row=1, column=1, value="SIMET / INCRA — Dados de Ofertas WEB Grandes Propriedades R$/ha")
        cel.font = Font(bold=True, color="FFFFFF", size=12)
        cel.fill = verde
        cel.alignment = center

        # Subtitulo com data
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(_COLUNAS_GRANDES))
        cel = ws.cell(row=2, column=1,
                      value=f"Data de Atualização: {datetime.now().strftime('%d/%m/%Y')}")
        cel.font = Font(italic=True, size=9)
        cel.alignment = Alignment(horizontal="left", vertical="center")

        # Cabecalho
        for idx, nome in enumerate(_COLUNAS_GRANDES, start=1):
            c = ws.cell(row=3, column=idx, value=nome)
            c.font = branco_bold
            c.fill = verde_claro
            c.alignment = center
            c.border = borda

        # Dados
        col_money = {6, 7, 9, 10, 12, 13, 15, 16, 18, 19}  # Mediana e Media de cada tipologia
        col_pct = {8, 11, 14, 17, 20}                       # CV de cada tipologia
        col_int = {4}                                       # COD. IBGE

        for ridx, row in enumerate(df.itertuples(index=False), start=4):
            for cidx, val in enumerate(row, start=1):
                if cidx in col_money:
                    cel = ws.cell(row=ridx, column=cidx, value=(float(val) if val is not None else None))
                    cel.number_format = '#,##0.00'
                elif cidx in col_pct:
                    cel = ws.cell(row=ridx, column=cidx, value=(float(val) if val is not None else None))
                    cel.number_format = '#,##0.00" %"'
                elif cidx in col_int:
                    cel = ws.cell(row=ridx, column=cidx, value=(int(val) if val is not None else None))
                else:
                    cel = ws.cell(row=ridx, column=cidx, value=val)
                cel.border = borda
                cel.alignment = Alignment(
                    horizontal="right" if cidx in (col_money | col_pct | col_int) else "left",
                    vertical="center",
                )

        ws.freeze_panes = "F4"  # congela ate Municipio + linha de header

        larguras = [14, 26, 6, 12, 28] + [16, 16, 11] * 5
        for i, w in enumerate(larguras, start=1):
            ws.column_dimensions[_gcl(i)].width = w

        ws.row_dimensions[1].height = 22
        ws.row_dimensions[3].height = 32

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        nome = f"simet_grandes_propriedades_{ts}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{nome}"'},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/relatorio/grandes-propriedades/pdf")
def exportar_grandes_pdf(estado: str, mercado_regional: str | None = None,
                         regiao: str | None = None, municipio: str | None = None,
                         user: dict = Depends(verificar_token)):
    """Relatorio PDF de Grandes Propriedades (fazendas >= 50 ha) — exige UF."""
    if not estado or estado.lower() == "todos":
        raise HTTPException(status_code=400, detail="Parametro 'estado' (UF) e obrigatorio no PDF.")

    try:
        from reportlab.lib.pagesizes import A3, landscape
        from reportlab.lib.colors import HexColor, black
        from reportlab.pdfgen import canvas
        from reportlab.platypus import Table, TableStyle
        from reportlab.lib import colors as rl_colors

        df = _buscar_dados_grandes(regiao=regiao, estado=estado,
                                   municipio=municipio, mercado_regional=mercado_regional)

        mercado_nome = ""
        if mercado_regional and mercado_regional.lower() != "todos" and not df.empty:
            nome_col = df["mercado_regional"].dropna()
            if not nome_col.empty:
                mercado_nome = str(nome_col.iloc[0])

        buffer = io.BytesIO()
        largura, altura = landscape(A3)
        c = canvas.Canvas(buffer, pagesize=landscape(A3))

        def cabecalho():
            _desenhar_logo_incra(c, 36, altura - 72, tamanho=52)
            c.setFillColor(HexColor("#3D6A1C"))
            c.setFont("Helvetica-Bold", 17)
            c.drawString(104, altura - 38, "SIMET / INCRA · Dados de Ofertas WEB Grandes Propriedades R$/ha")
            c.setFillColor(black)
            c.setFont("Helvetica", 10)
            sub = f"UF: {estado}"
            if mercado_nome:
                sub += f" · Mercado Regional: {mercado_nome}"
            sub += f" · Data de Atualização: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            c.drawString(104, altura - 58, sub)
            c.setStrokeColor(HexColor("#6FA030"))
            c.setLineWidth(1.5)
            c.line(36, altura - 84, largura - 36, altura - 84)

        def rodape(pagina, total):
            c.setFont("Helvetica-Oblique", 8)
            c.setFillColor(HexColor("#888888"))
            c.drawString(36, 20, "SIMET · Sistema de Inteligencia de Mercado de Terras · INCRA")
            c.drawRightString(largura - 36, 20, f"Pagina {pagina}/{total}")
            c.setFillColor(black)

        # Cabecalho hierarquico em 2 linhas
        # Linha 0: dimensoes (5) + 5 grupos de tipologia (3 colunas cada)
        super_header = ["Região", "Mercado Regional", "UF", "COD. IBGE", "Município",
                        "Geral", "", "", "Agrícola", "", "",
                        "Pecuária", "", "", "Floresta Plantada", "", "",
                        "Vegetação Nativa", "", ""]
        sub_header = ["", "", "", "", "",
                      "Mediana", "Média", "CV", "Mediana", "Média", "CV",
                      "Mediana", "Média", "CV", "Mediana", "Média", "CV",
                      "Mediana", "Média", "CV"]

        col_widths = [62, 100, 24, 50, 110,
                      52, 52, 38, 52, 52, 38, 52, 52, 38, 52, 52, 38, 52, 52, 38]

        if df.empty:
            cabecalho()
            c.setFont("Helvetica", 12)
            c.drawString(40, altura - 120, "Nenhum dado encontrado para os filtros selecionados.")
            rodape(1, 1)
            c.save()
        else:
            linhas_dados = []
            for _, r in df.iterrows():
                linhas_dados.append([
                    str(r.get("regiao") or ""),
                    str(r.get("mercado_regional") or "—"),
                    str(r.get("uf") or ""),
                    str(int(r["mun_cod"])) if r.get("mun_cod") is not None else "",
                    str(r.get("municipio") or ""),
                    _fmt_brl_num(r.get("mediana_geral")),
                    _fmt_brl_num(r.get("media_geral")),
                    _fmt_pct(r.get("coef_dispersao_geral")),
                    _fmt_brl_num(r.get("mediana_agricola")),
                    _fmt_brl_num(r.get("media_agricola")),
                    _fmt_pct(r.get("coef_dispersao_agricola")),
                    _fmt_brl_num(r.get("mediana_pecuaria")),
                    _fmt_brl_num(r.get("media_pecuaria")),
                    _fmt_pct(r.get("coef_dispersao_pecuaria")),
                    _fmt_brl_num(r.get("mediana_floresta_plantada")),
                    _fmt_brl_num(r.get("media_floresta_plantada")),
                    _fmt_pct(r.get("coef_dispersao_floresta_plantada")),
                    _fmt_brl_num(r.get("mediana_vegetacao_nativa")),
                    _fmt_brl_num(r.get("media_vegetacao_nativa")),
                    _fmt_pct(r.get("coef_dispersao_vegetacao_nativa")),
                ])

            chunk = 28  # linhas por pagina
            total_pags = max(1, (len(linhas_dados) + chunk - 1) // chunk)
            for p in range(total_pags):
                cabecalho()
                inicio = p * chunk
                fim = min(len(linhas_dados), inicio + chunk)
                trecho = [super_header, sub_header] + linhas_dados[inicio:fim]
                tbl = Table(trecho, colWidths=col_widths, repeatRows=2)
                tbl.setStyle(TableStyle([
                    # SPANs (precisam vir antes dos estilos das celulas mescladas)
                    ("SPAN", (0, 0), (0, 1)),
                    ("SPAN", (1, 0), (1, 1)),
                    ("SPAN", (2, 0), (2, 1)),
                    ("SPAN", (3, 0), (3, 1)),
                    ("SPAN", (4, 0), (4, 1)),
                    ("SPAN", (5, 0), (7, 0)),
                    ("SPAN", (8, 0), (10, 0)),
                    ("SPAN", (11, 0), (13, 0)),
                    ("SPAN", (14, 0), (16, 0)),
                    ("SPAN", (17, 0), (19, 0)),
                    # Dimensoes (cols 0..4, ambas as linhas) — verde escuro nas 2 linhas
                    ("BACKGROUND", (0, 0), (4, 1), HexColor("#3D6A1C")),
                    ("TEXTCOLOR", (0, 0), (4, 1), rl_colors.white),
                    ("FONTNAME", (0, 0), (4, 1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (4, 1), 8),
                    ("ALIGN", (0, 0), (4, 1), "CENTER"),
                    # Super-header das tipologias (cols 5..19, linha 0)
                    ("BACKGROUND", (5, 0), (-1, 0), HexColor("#3D6A1C")),
                    ("TEXTCOLOR", (5, 0), (-1, 0), rl_colors.white),
                    ("FONTNAME", (5, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (5, 0), (-1, 0), 8),
                    ("ALIGN", (5, 0), (-1, 0), "CENTER"),
                    # Sub-header das metricas (cols 5..19, linha 1)
                    ("BACKGROUND", (5, 1), (-1, 1), HexColor("#6FA030")),
                    ("TEXTCOLOR", (5, 1), (-1, 1), rl_colors.white),
                    ("FONTNAME", (5, 1), (-1, 1), "Helvetica-Bold"),
                    ("FONTSIZE", (5, 1), (-1, 1), 7),
                    ("ALIGN", (5, 1), (-1, 1), "CENTER"),
                    # Corpo
                    ("FONTSIZE", (0, 2), (-1, -1), 6.5),
                    ("ROWBACKGROUNDS", (0, 2), (-1, -1), [rl_colors.whitesmoke, rl_colors.white]),
                    ("ALIGN", (2, 2), (2, -1), "CENTER"),     # UF
                    ("ALIGN", (3, 2), (3, -1), "CENTER"),     # COD. IBGE
                    ("ALIGN", (5, 2), (-1, -1), "RIGHT"),     # numericos
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, 1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, 1), 6),
                    ("TOPPADDING", (0, 2), (-1, -1), 2),
                    ("BOTTOMPADDING", (0, 2), (-1, -1), 2),
                    ("GRID", (0, 0), (-1, -1), 0.25, HexColor("#CCCCCC")),
                ]))
                w, h = tbl.wrapOn(c, largura - 72, altura - 160)
                tbl.drawOn(c, 36, altura - 100 - h)
                rodape(p + 1, total_pags)
                if p + 1 < total_pags:
                    c.showPage()
            c.save()

        buffer.seek(0)
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        nome = f"simet_grandes_propriedades_{estado}_{ts}.pdf"
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{nome}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# =====================================================================
# GESTAO DE USUARIOS (nivel 0 apenas, exceto /me)
# =====================================================================
class NovoUsuarioReq(BaseModel):
    usuario: str
    senha: str
    nivel: int

class AtualizarUsuarioReq(BaseModel):
    nivel: int | None = None
    ativo: bool | None = None

class ResetSenhaReq(BaseModel):
    nova_senha: str


@app.get("/api/me")
def obter_perfil_atual(user: dict = Depends(verificar_token)):
    return {"usuario": user["username"], "nivel": user["nivel"], "id": user.get("id")}


@app.get("/api/saude")
def saude_base(user: dict = Depends(verificar_token)):
    """Retorna o snapshot de saude da base (validos + contagens por status
    da fila de processamento). Usado na Central de Comando para acompanhar
    o backlog sem precisar executar uma operacao."""
    conn = obter_conexao()
    try:
        return {"sucesso": True, **contagem_saude_fila(conn)}
    finally:
        conn.close()


@app.get("/api/usuarios")
def listar_usuarios(user: dict = Depends(exigir_nivel(0))):
    return {"sucesso": True, "usuarios": usuarios_db.listar()}


@app.post("/api/usuarios")
def criar_usuario(req: NovoUsuarioReq, user: dict = Depends(exigir_nivel(0))):
    try:
        novo_id = usuarios_db.criar(req.usuario.strip(), req.senha, int(req.nivel))
        return {"sucesso": True, "id": novo_id}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Nao foi possivel criar usuario: {e}")


@app.patch("/api/usuarios/{usr_id}")
def atualizar_usuario(usr_id: int, req: AtualizarUsuarioReq,
                      user: dict = Depends(exigir_nivel(0))):
    if usr_id == user.get("id"):
        if req.ativo is False:
            raise HTTPException(status_code=400, detail="Voce nao pode desativar a si mesmo.")
        if req.nivel is not None and req.nivel != 0:
            raise HTTPException(status_code=400, detail="Voce nao pode rebaixar a si mesmo.")

    if (req.nivel is not None and req.nivel != 0) or req.ativo is False:
        alvo = usuarios_db.buscar_por_id(usr_id)
        if alvo and alvo["nivel"] == 0 and alvo["ativo"]:
            if usuarios_db.contar_admins_ativos() <= 1:
                raise HTTPException(
                    status_code=400,
                    detail="Nao e possivel remover o poder do ultimo administrador ativo.",
                )

    try:
        usuarios_db.atualizar(usr_id, nivel=req.nivel, ativo=req.ativo)
        return {"sucesso": True}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/usuarios/{usr_id}/reset-senha")
def resetar_senha_usuario(usr_id: int, req: ResetSenhaReq,
                          user: dict = Depends(exigir_nivel(0))):
    try:
        usuarios_db.resetar_senha(usr_id, req.nova_senha)
        return {"sucesso": True}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.delete("/api/usuarios/{usr_id}")
def excluir_usuario(usr_id: int, user: dict = Depends(exigir_nivel(0))):
    if usr_id == user.get("id"):
        raise HTTPException(status_code=400, detail="Voce nao pode excluir a si mesmo.")
    alvo = usuarios_db.buscar_por_id(usr_id)
    if alvo and alvo["nivel"] == 0 and alvo["ativo"]:
        if usuarios_db.contar_admins_ativos() <= 1:
            raise HTTPException(
                status_code=400,
                detail="Nao e possivel excluir o ultimo administrador ativo.",
            )
    usuarios_db.excluir(usr_id)
    return {"sucesso": True}


# =====================================================================
# SERVIR O FRONTEND BUILDADO (modo "executavel")
# =====================================================================
# Quando frontend_react/dist/ existe (gerado por `npm run build`), o uvicorn
# passa a servir o app inteiro num unico processo + porta. Sem o build, esse
# bloco e no-op e o front segue rodando via `npm run dev` em outra porta.
# IMPORTANTE: mount fica POR ULTIMO para nao sobrepor as rotas /api/*.
_FRONT_DIST = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           "frontend_react", "dist")

if os.path.isdir(_FRONT_DIST):
    # Assets versionados pelo Vite (hash no nome) — cache longo seguro
    _ASSETS_DIR = os.path.join(_FRONT_DIST, "assets")
    if os.path.isdir(_ASSETS_DIR):
        app.mount("/assets", StaticFiles(directory=_ASSETS_DIR), name="assets")

    _INDEX_HTML = os.path.join(_FRONT_DIST, "index.html")

    # index.html NAO pode cachear — quando o build muda os hashes dos bundles
    # mudam, mas o navegador precisa pegar o index.html novo pra saber disso.
    # Os bundles JS/CSS sao seguros pra cache longo (URL versionada).
    _NO_CACHE = {"Cache-Control": "no-cache, no-store, must-revalidate"}

    @app.get("/{caminho:path}", include_in_schema=False)
    def _spa_catch_all(caminho: str, request: Request):
        # Bloqueia rotas de API que nao casaram (404 explicito ao inves de
        # devolver index.html, que confundiria fetches do front).
        if caminho.startswith("api/") or caminho.startswith("api"):
            raise HTTPException(status_code=404, detail="Endpoint nao encontrado")

        # Tenta servir um arquivo estatico real (favicon, robots.txt, etc.)
        candidato = os.path.join(_FRONT_DIST, caminho) if caminho else _INDEX_HTML
        if caminho and os.path.isfile(candidato):
            # index.html (caso raro de ser solicitado direto): no-cache
            if caminho == "index.html":
                return FileResponse(candidato, headers=_NO_CACHE)
            return FileResponse(candidato)
        # SPA fallback: qualquer outra rota devolve o index.html (sem cache)
        return FileResponse(_INDEX_HTML, headers=_NO_CACHE)

    print(f"[STATIC] Front buildado servido de {_FRONT_DIST}", flush=True)
else:
    print(f"[STATIC] {_FRONT_DIST} nao existe — front em modo dev (rode `npm run dev`)",
          flush=True)